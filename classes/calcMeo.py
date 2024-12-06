from datetime import datetime
import inspect
import numpy as np
import os
import pandas as pd
import sqlalchemy
import sqlite3
import subprocess
import sys
import time
import traceback

from pymeo.classes.apiAna import hidroWeb
from pymeo.config import PATH, PATHARQUIVOS

# from pymeo.classes.apiAna import hidroWeb
#import utilsPymeo
# dirClasses = os.path.abspath(os.path.join(os.path.dirname(__file__), '../classes'))
# sys.path.insert(0, dirClasses)
# from apiAna import hidroWeb

dirComuns = os.path.abspath(os.path.join(os.path.dirname(__file__), '../comuns'))
sys.path.insert(0, dirComuns)
import utils
import utilsPymeo

"""

 
Autores
Diego Pinheiro de Menezes
Felipe (Zago)
Regina Celia Enedina Pereira Gomes da Silva

"""


class meo:
    def __init__(self, tipo):
        """
        Classe criada para realizar o calculo da Média das Enchentes Ordinárias (MEO)
        
        params
        tipo
        0 - Via API Hidro Webservice
        1 - Via arquivo csv

        ...
        Attributes
        ----------
        codEstacao : string
            Código da estação que será efetuado o cálculo
        nivel : int
            Nível de consistência das séries históricas.
            Default: None

            None - Incluir todos os níveis no cálculo
            1 - Bruto
            2 - Consistido (Já foi validado pela ANA)


        Methods
        -------
        totalReg
        calcMeo(df, nivel)

        Parâmetros
        :param
        df: DataFrame com os dados que serão utilizados para realizar o cálculo
        Os dados inseridos no df de origem, pode vir da API da Ana ou do Portal Hidro Web (mdb ou txt), desde que,
        as colunas do df estejam no mesmo formato dos dados obtidos pelo endpoint HidroSerieCotas

        nivel: Nível de consistência 1 ou 2
        1 - Bruto
        2 - Consistido (Já foi validado pela ANA)
        """
        
        if tipo==0:
            self.objHidro = hidroWeb()
            self.objHidro.menorData = None
            self.objHidro.df = pd.DataFrame()       # Dataframe que traz o resultado dos dados do objHidro (Classe hidroWeb)
        
        self.titulo   = "Cálculo da Meo"

        # Dados que deverão ser definidos via arquivo de configuração para que o usuário tenha a opção de alterar
        self.exibirMsg = False

        # Valor padrão das variáveis
        self.resultado  = -1
        self.codEstacao = None
        self.nivel      = 2
        self.dtInicio   = None
        self.dtFim      = None
        self.origem     = None
        self.arqCsv     = None
        self.tipoDados = 'Via Api Hidro Webservice'
        self.totalRegSerieNivel = 0         #Total de registros após filtrar o dataframe pelo nível de consistência selecionado
        self.numAnos       = 0
        self.totalRegSerie = 0
        
        self.nomeEstacao = None
        self.codRio      = None
        self.nomeRio     = None
        self.codMunicipio = None
        self.nomeMunicipio = None
        self.siglaResponsavel = None
        
        # Inicializar os DFs que serão usados na classe
        self.df          = pd.DataFrame()       # Dataframe que traz o resultado dos dados da classe do módulo vigente (Classe meo)
        df               = pd.DataFrame()       # Dataframe que armazena dados da função local em que está sendo utilizada
        
        # Dataframes que podem ser usados fora da classe
        self.dfEstacao   = pd.DataFrame()
        self.dfResultado = pd.DataFrame()
        self.dfFinal     = pd.DataFrame()
        self.dfMeoDet    = pd.DataFrame()
        self.dfMeo       = pd.DataFrame()
        self.dfSerieCompleta = pd.DataFrame()

        self.dfMaximasAnuais = None

    # Informar o arqCsv para a opção de Cálculo da Meo usando o arquivo Csv obtido via Portal Hidroweb
    def calcularMeo(self, codEstacao, nivel=None, dtInicio=None, dtFim=None, origem=None, arqCsv=None):    
        try:
            self.codEstacao = codEstacao

            if nivel:
               self.nivel = nivel
            
            if dtInicio:
                self.dtInicio = dtInicio
                
            if dtFim:
                self.dtFim = dtFim
            
            if origem:
                self.origem = origem
            
            if arqCsv:
                self.arqCsv = arqCsv
                self.calculoViaCsv()
            else:
                self.arqCsv = None
                self.calculoViaApi()
                
        except Exception as e:
            erro = str(e)
            print(erro)
            utilsPymeo.gerarLog(2,
                            [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                            f"Problema na execução da função calcularMeo",
                            # f"Status: {self.objHidro.status}",
                            #f"Estação: {self.codEstacao}",
                            #f"Registros no Df: {len(self.objHidro.df)}",
                            erro
                            ], 
                            inspect.getframeinfo(inspect.currentframe()), 
                            self.exibirMsg)
            #sys.exit(1)

    ##############################################################################################
    # Calcular a Meo a partir dos dados obtidos via arquivo csv 
    def calculoViaCsv(self):
        self.obterDadosSerieViaCsv()
        self.prepararDadosObtidosViaCsv()
        self.calcular()

    def obterDadosEstacaoViaCsv(self):
        # Busca dos dados da estação
        try:
            # dirDados = os.path.abspath(os.path.join(os.path.dirname(__file__)))
            dirDados = os.path.abspath(os.path.join(os.path.dirname(__file__), '../dados'))
            sgdb = 'sqlite'
            db = f"{dirDados}\\pymeo.db"
            engine = sqlalchemy.create_engine(f"{sgdb}:///{db}")
            cnx = sqlite3.connect(f"{db}")
            cursor = cnx.cursor()
        
            sql = f"SELECT * FROM tb_estacao WHERE codigoEstacao = {self.codEstacao}"
            
            self.dfEstacao = pd.read_sql_query(sql, cnx)

            self.nomeEstacao = self.dfEstacao["nomeEstacao"].iloc[0]
            self.codRio = self.dfEstacao['codigoRio'].iloc[0]
            self.nomeRio = self.dfEstacao['nomeRio'].iloc[0]
            self.codMunicipio = self.dfEstacao['codigoMunicipio'].iloc[0]
            self.nomeMunicipio = self.dfEstacao['nomeMunicipio'].iloc[0]
            self.siglaResponsavel = self.dfEstacao['siglaResponsavel'].iloc[0]
            self.siglaOperadora = self.dfEstacao['siglaOperadora'].iloc[0]

            # Retorna a menor data obtida entre uma lista de datas obtidos via Api
            #if self.dtInicio is None:
            #    self.dtInicio = self.objHidro.menorData
            #
            #if self.dtFim is None:
            #    self.dtFim    = datetime.now().strftime("%Y/%m/%d")
                
        except Exception as e:
            self.resultado = -1
            erro = str(e)
            utilsPymeo.gerarLog(2,
                            [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                            f"Problema na execução da obterDadosEstacaoViaApi",
                            # f"Status: {self.objHidro.status}",
                            #f"Estação: {self.codEstacao}",
                            #f"Registros no Df: {len(self.objHidro.df)}",
                            erro],
                            inspect.getframeinfo(inspect.currentframe()),
                            self.exibirMsg)
            #sys.exit(1)
            raise

    def obterDadosSerieViaCsv(self):
        # Busca dos dados da série, sem realizar o cálculo

        try:
            self.dfFinal = pd.read_csv(self.arqCsv, sep=';', encoding='ANSI', engine='c', skiprows=15)
            
            self.obterDadosEstacaoViaCsv()

            # Dataframe obtido via Api que será utilizado para armazenar os dados da série
            self.dfSerieCompleta = self.dfFinal
            self.df              = self.dfFinal
        except Exception as e:
            self.resultado = -2
            erro = str(e)
            utilsPymeo.gerarLog(2,
                            [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                            f"Problema na execução da obterDadosSerieViaCsv",
                            #f"Estação: {self.codEstacao}",
                            #f"Registros no Df: {len(self.objHidro.df)}",
                            erro], 
                            inspect.getframeinfo(inspect.currentframe()), 
                            self.exibirMsg)
            #sys.exit(1)
            raise

    def prepararDadosObtidosViaCsv(self):
        try:
            totreg = len(self.df)
            nivel = str(self.nivel)
            self.totalRegSerie = len(self.df)
            # Adiciona a coluna 'AnoMes' baseada na coluna 'Data'
            self.df['Ano'] = self.df['Data'].apply(lambda x: str(x)[6:10])
            self.df['Mes'] = self.df['Data'].apply(lambda x: str(x)[3:5])
            self.df['AnoMes'] = self.df['Ano']+self.df['Mes']
            
            self.df.rename(columns={'NivelConsistencia': 'nivelconsistencia', 'DiaMaxima': 'Dia_Maxima', 'DiaMinima': 'Dia_Minima', 'MediaDiaria': 'Mediadiaria'}, inplace=True)
            self.df['Mediadiaria'] = self.df['Mediadiaria'].astype(str)
            self.df['nivelconsistencia'] = self.df['nivelconsistencia'].astype(str)
    
            # Filtragem do dataframe com base nos demais critérios (Nivel de Consistência 2, Média Diária 1-sim, e Máximas sem valores em branco)
            if nivel == '0':
                # 0 - Todos os níveis de consistência
                dfFiltrado = self.df[(self.df['Mediadiaria'] == '1') & (~self.df['Maxima'].isna())].copy()
            else:
                # Filtrar pelo nível de consistência informado
                # 1 - Bruto
                # 2 - Consistido
                dfFiltrado = self.df[(self.df['Mediadiaria'] == '1') & (~self.df['Maxima'].isna()) & (self.df['nivelconsistencia'] == nivel)].copy()  # uso de cópia para não alterar arquivo original
            self.totalRegSerieNivel = len(dfFiltrado)
            
            # Convertendo DiaMaxima em String
            dfFiltrado['Dia_Maxima'] = dfFiltrado['Dia_Maxima'].astype('object')
            # Aplica uma função lambda à coluna 'DiaMaxima' que, se o valor não for nulo (pd.notnull(x)), formata o número do dia
            dfFiltrado['Dia_Maxima'] = dfFiltrado['Dia_Maxima'].apply(lambda x: f"{int(x):02d}" if pd.notnull(x) else '')
            # Cria uma nova coluna 'DataMaxima' no DataFrame 'dfFiltrado', que contém a data formatada no estilo brasileiro de dia/mês/ano como uma string.
            dfFiltrado['DataMaxima'] = dfFiltrado['Dia_Maxima'] + '/' + dfFiltrado['Mes'] + '/' + dfFiltrado['Ano']
            # Convertendo DataMaxima em Datetime
            dfFiltrado['DataMaxima'] = pd.to_datetime(dfFiltrado['DataMaxima'], format='%d/%m/%Y')
            # Extraia apenas a parte da data, descartando a hora, se houver
            dfFiltrado['DataMaxima'] = dfFiltrado['DataMaxima'].dt.date
    
            # Substituir os valores brancos e nulos para zero e converter as colunas para inteiro
            # .replace('', 0) - Substituir os valores em branco por 0
            # .fillna(0)      - Substituir os valores nulos por 0
            # .astype(int)    - Converter a coluna para inteiro 
            dfFiltrado['Maxima'] = dfFiltrado.Maxima.replace('', 0).fillna(0).astype(float).astype(int)
            dfFiltrado['Minima'] = dfFiltrado.Minima.replace('', 0).fillna(0).astype(float).astype(int)
            dfFiltrado['Media']  = dfFiltrado.Media.replace('', 0).fillna(0).astype(float).astype(int)
    
            self.dfFinal = dfFiltrado[
                ['nivelconsistencia', 'Maxima', 'Minima', 'Media', 'Dia_Maxima', 'Dia_Minima', 'Mes', 'Ano', 'DataMaxima']]
            self.gerarDfMeoDetalhes(self.dfFinal)
        except Exception as e:
            if totreg>0:
                self.resultado = -3
                self.salvarDfToCsv(self.df, f"E:\\pymeo\\arquivos\\{self.codEstacao}_df.csv")
            else:
                self.resultado = -4
                
                erro = str(e)
                utilsPymeo.gerarLog(2,
                                [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                                f"Problema na execução da função calcularMeo",
                                # f"Status: {self.objHidro.status}",
                                #f"Estação: {self.codEstacao}",
                                #f"Registros no Df: {len(self.objHidro.df)}",
                                erro
                                ], 
                                inspect.getframeinfo(inspect.currentframe()), 
                                self.exibirMsg)
                #sys.exit(1)
                raise
                
    ##############################################################################################                
    ##############################################################################################
    # Calcular a Meo a partir dos dados obtidos via API Hidro Webservice
    def calculoViaApi(self):
        self.obterDadosSerieViaApi()
        self.prepararDadosObtidosViaApi()
        self.calcular()
    
    
    def obterDadosEstacaoViaApi(self):
        # Busca dos dados da estação
        try:
            self.objHidro.getDadosHidroInventarioEstacoes(self.codEstacao)

            # Retorna a menor data obtida entre uma lista de datas obtidos via Api
            if self.dtInicio is None:
                self.dtInicio = self.objHidro.menorData

            if self.dtFim is None:
                self.dtFim    = datetime.now().strftime("%Y/%m/%d")

            # Dataframe obtido via Api que será utilizado para armazenar os dados da estação
            self.dfEstacao = self.objHidro.dfEstacao
            
            self.nomeEstacao = self.dfEstacao["Estacao_Nome"].values[0]
            self.codRio = self.dfEstacao['Rio_Codigo'].values[0]
            self.nomeRio = self.dfEstacao['Rio_Nome'].values[0]
            self.codMunicipio = self.dfEstacao['Municipio_Codigo'].values[0]
            self.nomeMunicipio = self.dfEstacao['Municipio_Nome'].values[0]
            self.siglaResponsavel = self.dfEstacao['Responsavel_Sigla'].values[0]
            
            # O objeto self.df é criado por padrão após a chamada de uma função da API
            # Avaliar se é necessário
            self.df        = self.objHidro.df
        except Exception as e:
            self.resultado = -1
            erro = str(e)
            utilsPymeo.gerarLog(2,
                            [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                            f"Problema na execução da obterDadosEstacaoViaApi",
                            f"Status: {self.objHidro.status}",
                            #f"Estação: {self.codEstacao}",
                            #f"Registros no Df: {len(self.objHidro.df)}",
                            erro],
                            inspect.getframeinfo(inspect.currentframe()),
                            self.exibirMsg)
            #sys.exit(1)
            raise

    def obterDadosSerieViaApi(self):
        # Busca dos dados da série, sem realizar o cálculo

        try:
            # É obrigatório buscar os dados da estação antes de buscar as séries
            self.obterDadosEstacaoViaApi()
                        
            # A função obterDadosEstacaoViaApi gera os valores das variáveis self.dtInicio e self.dtFim, eles serão substituídos apenas se o usuário passar nos parâmetros dtInicio e dtFim que são opcionais
            #if self.dtInicio:
            #    self.dtInicio = dtInicio

            #if self.dtFim:
            #    self.dtFim = dtFim

            self.thread = self.objHidro.getDadosHidroSerieCotas(self.dtInicio, self.dtFim, self.exibirMsg, origem=self.origem)
                
            # Dataframe obtido via Api que será utilizado para armazenar os dados da série
            self.dfSerieCompleta = self.objHidro.df
            self.df              = self.objHidro.df
        except Exception as e:
            self.resultado = -2
            erro = str(e)
            utilsPymeo.gerarLog(2,
                            [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                            f"Problema na execução da obterDadosSerieViaApi",
                            f"Status: {self.objHidro.status}",
                            #f"Estação: {self.codEstacao}",
                            #f"Registros no Df: {len(self.objHidro.df)}",
                            erro], 
                            inspect.getframeinfo(inspect.currentframe()), 
                            self.exibirMsg)
            #sys.exit(1)
            raise
    
              
    def prepararDadosObtidosViaApi(self):
        try:
            totreg = len(self.df)
            nivel = str(self.nivel)
            self.totalRegSerie = len(self.df)
            # Adiciona a coluna 'AnoMes' baseada na coluna 'Data'
            self.df['AnoMes'] = self.df['Data_Hora_Dado'].apply(lambda x: str(x)[0:10])
    
            # Filtragem do dataframe com base nos demais critérios (Nivel de Consistência 2, Média Diária 1-sim, e Máximas sem valores em branco)
            if nivel == '0':
                # 0 - Todos os níveis de consistência
                dfFiltrado = self.df[(self.df['Mediadiaria'] == '1') & (~self.df['Maxima'].isna())].copy()
            else:
                # Filtrar pelo nível de consistência informado
                # 1 - Bruto
                # 2 - Consistido
                dfFiltrado = self.df[(self.df['Mediadiaria'] == '1') & (~self.df['Maxima'].isna()) & (self.df['nivelconsistencia'] == nivel)].copy()  # uso de cópia para não alterar arquivo original
            self.totalRegSerieNivel = len(dfFiltrado)
    
            # Cria uma nova coluna 'Mes' no DataFrame 'dfFiltrado', que contém o mês extraído da coluna 'Data'.
            dfFiltrado['Mes'] = dfFiltrado['Data_Hora_Dado'].apply(
                lambda x: str(x)[5:7])
            # Cria uma nova coluna 'Ano' no DataFrame 'dfFiltrado', que contém o ano extraído da coluna 'Data'.
            dfFiltrado['Ano'] = dfFiltrado['Data_Hora_Dado'].apply(lambda x: str(x)[0:4])
    
            # Convertendo DiaMaxima em String
            dfFiltrado['Dia_Maxima'] = dfFiltrado['Dia_Maxima'].astype('object')
            # Aplica uma função lambda à coluna 'DiaMaxima' que, se o valor não for nulo (pd.notnull(x)), formata o número do dia
            dfFiltrado['Dia_Maxima'] = dfFiltrado['Dia_Maxima'].apply(lambda x: f"{int(x):02d}" if pd.notnull(x) else '')
            # Cria uma nova coluna 'DataMaxima' no DataFrame 'dfFiltrado', que contém a data formatada no estilo brasileiro de dia/mês/ano como uma string.
            dfFiltrado['DataMaxima'] = dfFiltrado['Dia_Maxima'] + '/' + dfFiltrado['Mes'] + '/' + dfFiltrado['Ano']
            # Convertendo DataMaxima em Datetime
            dfFiltrado['DataMaxima'] = pd.to_datetime(dfFiltrado['DataMaxima'], format='%d/%m/%Y')
            # Extraia apenas a parte da data, descartando a hora, se houver
            dfFiltrado['DataMaxima'] = dfFiltrado['DataMaxima'].dt.date
    
            # Substituir os valores brancos e nulos para zero e converter as colunas para inteiro
            # .replace('', 0) - Substituir os valores em branco por 0
            # .fillna(0)      - Substituir os valores nulos por 0
            # .astype(int)    - Converter a coluna para inteiro 
            dfFiltrado['Maxima'] = dfFiltrado.Maxima.replace('', 0).fillna(0).astype(float).astype(int)
            dfFiltrado['Minima'] = dfFiltrado.Minima.replace('', 0).fillna(0).astype(float).astype(int)
            dfFiltrado['Media']  = dfFiltrado.Media.replace('', 0).fillna(0).astype(float).astype(int)
    
            self.dfFinal = dfFiltrado[
                ['nivelconsistencia', 'Maxima', 'Minima', 'Media', 'Dia_Maxima', 'Dia_Minima', 'Mes', 'Ano', 'DataMaxima']]
            self.gerarDfMeoDetalhes(self.dfFinal)
        except Exception as e:
            if totreg>0:
                self.resultado = -3
                self.salvarDfToCsv(self.df, f"E:\\pymeo\\arquivos\\{self.codEstacao}_df.csv")
            else:
                self.resultado = -4
                
                erro = str(e)
                utilsPymeo.gerarLog(2,
                                [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                                f"Problema na execução da função calcularMeo",
                                f"Status: {self.objHidro.status}",
                                #f"Estação: {self.codEstacao}",
                                #f"Registros no Df: {len(self.objHidro.df)}",
                                erro
                                ], 
                                inspect.getframeinfo(inspect.currentframe()), 
                                self.exibirMsg)
                #sys.exit(1)
                raise

    ##############################################################################################
    # Calcula o MEO (Média das Enchentes Ordinárias) com base nos dados obtidos via Api (prepararDadosObtidosViaApi) ou CSV
    def calcular(self):
        try:
            totreg = len(self.dfFinal)

            # Verifica se há dados suficientes para o cálculo (pelo menos 20 anos).
            self.numAnos = self.dfFinal['Ano'].nunique()
            if self.numAnos < 20:
                return None

            self.maiorData = self.dfFinal['DataMaxima'].max()
            self.menorData = self.dfFinal['DataMaxima'].min()
            

            # Conta o número de anos únicos no conjunto de dados.
            self.numAnosUnico = self.dfFinal['Ano'].nunique()

            # Agrupando por 'Ano' e obtendo o valor máximo de 'Maxima' para cada ano
            self.dfMaximasAnuais = self.dfFinal.groupby('Ano').agg({'Maxima': 'max', 'DataMaxima': 'first'}).reset_index()
            self.totalRegMaximasAnuais = len(self.dfMaximasAnuais)

            # Listagem das máximas em ordem decrescente para o Data Frame
            self.dfMaximasAnuais_ordenado = self.dfMaximasAnuais.sort_values(by='Maxima', ascending=False).reset_index(drop=True)

            # Equação para o cálculo da MEO
            # Computar o N da Fórmula P = 1/(n/N)
            N = len(self.dfMaximasAnuais_ordenado)
            # Computar o n para cada linha (que é simplesmente seu índice + 1)
            self.dfMaximasAnuais_ordenado['n'] = self.dfMaximasAnuais_ordenado.index + 1
            # Calcular o P para cada linha
            self.dfMaximasAnuais_ordenado['P'] = N / (self.dfMaximasAnuais_ordenado.index + 1)
            self.totalRegMaximasAnuaisOrdenado = len(self.dfMaximasAnuais_ordenado)

            # Calculando a Média das Enchentes Ordinárias (MEO)
            # print(self.dfMaximasAnuais_ordenado.columns.values)
            # Alterado em 29/10/2024, seguindo orientações do Diego
            #self.resultado = self.dfMaximasAnuais_ordenado.loc[(self.dfMaximasAnuais_ordenado['P'] >= 3.0001) & (self.dfMaximasAnuais_ordenado['P'] < 20), 'Maxima'].mean()
            self.resultado = self.dfMaximasAnuais_ordenado.loc[(self.dfMaximasAnuais_ordenado['P'] >= 2.99999) & (self.dfMaximasAnuais_ordenado['P'] < 20), 'Maxima'].mean()
            self.gerarMemoriaCalculo(self.codEstacao, self.nivel, self.resultado)
            #self.salvarDfToExcel(
            #        self.dfMaximasAnuais_ordenado,
            #        f"calcular_calculo_{self.codEstacao}.xlsx",
            #        planilha=self.codEstacao,
            #        titulo=f"Memória de cálculo da Estação {self.codEstacao}")

            self.gerarDfMeo()
        except Exception as e:
            if totreg>0:
                self.resultado = -5
                self.salvarDfToCsv(self.df, f"E:\\pymeo\\arquivos\\{self.codEstacao}_dffinal.csv")
            else:
                self.resultado = -6
            erro = str(e)
            print(erro)
            utilsPymeo.gerarLog(2,
                            [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                            f"Problema na execução da função calcular",
                            f"Status: {self.objHidro.status}",
                            #f"Estação: {self.codEstacao}",
                            #f"Registros no Df: {len(self.objHidro.df)}",
                            erro
                            ], 
                            inspect.getframeinfo(inspect.currentframe()), 
                            self.exibirMsg)
            #sys.exit(1)
            raise

    #def totalReg(self,df):
    #    self.totalReg = len(df)

    def gerarMemoriaCalculo(self, codEstacao, nivel, valorMeo):
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.formatting.rule import CellIsRule
        except ImportError:
            # Instalar o módulo caso não esteja
            dirPython = f"{sys.prefix}\\python.exe"
            subprocess.check_call([dirPython, "-m", "pip", "install", "openpyxl"])

        dtAtual = datetime.now().strftime('%Y%m%d')
        
                
        # Memória do cálculo
        df = self.dfMaximasAnuais_ordenado
        
        df['Nível'] = nivel
        
        # dirArquivos = os.path.abspath(os.path.join(os.path.dirname(__file__), '..\\arquivos'))
        
        # arquivo = f"{dirArquivos}\\{codEstacao}_{dtAtual}_{utils.limparString(self.nomeRio)}.xlsx"
        arquivo = f"{PATHARQUIVOS}\\{codEstacao}_{dtAtual}_{utils.limparString(self.nomeRio)}.xlsx"
        titulo =f"Memória de cálculo da Estação {codEstacao}"

        self.salvarDfToExcel(
                df,
                arquivo,
                planilha=str(codEstacao),
                titulo=titulo)
        
        arqExcel = load_workbook(arquivo)
        planilha = arqExcel.active
        
        # Adicionar duas linhas antes
        planilha.insert_rows(1, amount=11)
        
        # Inserir conteúdo nas linhas adicionadas
        # Linha 1
        planilha['A1'] = f"Estação "
        planilha['B1'] = f"{codEstacao}"
        planilha['C1'] = f"{self.nomeEstacao}"
        
        # Linha 2
        planilha['A2'] = f"* Estação operada por {self.siglaOperadora} com responsabilidade atribuída à {self.siglaResponsavel}"
        planilha['A2'].font = Font(size=8, bold=True, italic=True, color="FF0000")
        planilha.merge_cells(f"A2:F2")
        
        # Linha 3 - Em branco
        
        # Linha 4
        planilha['A4'] = f"Rio "
        planilha['B4'] = f"{self.codRio}"
        planilha['C4'] = f"{self.nomeRio}"
        
        # Linha 5
        planilha['A5'] = f"Município "
        planilha['B5'] = f"{self.codMunicipio}"
        planilha['C5'] = f"{self.nomeMunicipio}"
        
        # Linha 6 - Em branco
        
        # Linha 7
        planilha['A7'] = f"Valor da MEO"
        planilha['B7'] = f"{valorMeo}"
        
        # Linha 8
        planilha['A8'] = f"* O Valor da MEO foi calculado através dos dados obtidos {self.origem}"
        planilha['A8'].font = Font(size=8, bold=True, italic=True, color="FF0000")
        planilha.merge_cells(f"A8:F8")
        
        
        # Inserir 10 linhas na planilha a partir da última linha
        ultimaLinha = planilha.max_row
        planilha.insert_rows(ultimaLinha+3, amount=10)
        

        # O cálculo da MEO é ...
        # Inserir o texto a partir da 3ª linha depois da última linha
        linhaSobreCalculo = ultimaLinha + 3
        planilha['A{}'.format(linhaSobreCalculo)] = "O cálculo da MEO é uma estimativa da probabilidade da não excedência da vazão ordinária de um rio tendo como base uma série temporal mínima de 20 anos de observação das cotas ordenadas (do maior para a menor) de espelho d’água de uma dada estação fluviométrica, que pode ser sintetizada pela equação:"
        planilha.merge_cells(f"A{linhaSobreCalculo}:F{linhaSobreCalculo}")
        #planilha.row_dimensions[linhaSobreCalculo].height = 98
        planilha.row_dimensions[linhaSobreCalculo].height = 98
        planilha['A{}'.format(linhaSobreCalculo)].alignment = Alignment(horizontal='justify', vertical='center', wrap_text=True)
       
        # P = 1/(n/N)
        # Inserir o texto a partir da 4ª linha depois da última linha
        planilha['A{}'.format(ultimaLinha + 4)] = "P = 1/(n/N)"
        
        # Onde: 
        # Inserir o texto a partir da 5ª linha depois da última linha
        planilha['A{}'.format(ultimaLinha + 6)] = "Onde:"
        planilha['A{}'.format(ultimaLinha + 6)].font = Font(size=12, bold=True, italic=True, color="FF0000")
        
        # n = número da ...
        # Inserir o texto a partir da 6ª linha depois da última linha
        planilha['A{}'.format(ultimaLinha + 8)] = "n = número da vazão máxima na classificação em ordem decrescente (posição na lista)."
        planilha.merge_cells("A{}:F{}".format(ultimaLinha + 8, ultimaLinha + 8))
        #planilha.row_dimensions[ultimaLinha + 8].height = 36
        planilha.row_dimensions[ultimaLinha + 8].height = 36
        planilha['A{}'.format(ultimaLinha + 8)].alignment = Alignment(horizontal='justify', vertical='center', wrap_text=True)

        # N = número total de ...
        # Inserir o texto a partir da 7ª linha depois da última linha
        planilha['A{}'.format(ultimaLinha + 7)] = "N = número total de anos observados (tamanho da lista)."
        
        
        ####################################################################################################
        # Aplicar formatação condicional para o valor de P, que está na coluna E
        # Definir a formatação condicional da coluna E (P)
        formatoP = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Aplicar formatação condicional na coluna E (P)
        # A regra "between" irá aplicar a cor amarela nas células cujo valor esteja entre 3 e 20
        regra = CellIsRule(operator='between', formula=['3', '20'], fill=formatoP)
        
        # Selecionar as células da coluna E (P)
        celulas = 'E9:E{}'.format(planilha.max_row)
        
        # Aplicar a regra à coluna E
        planilha.conditional_formatting.add(celulas, regra)
        
               
        # Ajusta a largura da coluna e a altura da linha para visualizar melhor o texto
        # planilha.column_dimensions['A'].width = 50  # largura da coluna
        # planilha.row_dimensions[1].height = 40      # altura da linha
        
        arqExcel.save(arquivo)

    def salvarDfToCsv(self, df, arquivo):
        df.to_csv(arquivo, sep=';', header=True, )
    
    def salvarDfToPdf(self,df,arquivo,titulo=None):
        utilsPymeo.dfToPdf(df, arquivo, titulo)

    def salvarDfToExcel(self,df,arquivo,planilha=None,titulo=None):
        df.to_excel(arquivo, sheet_name=planilha, index=False)
        #print(os.path.abspath(arquivo))

    def gerarDfMeoDetalhes(self, df):
        self.dfMeoDet = pd.DataFrame()
        self.dfMeoDet = df.copy()
        self.dfMeoDet.insert(0, 'codigoEstacao', self.codEstacao)
        self.dfMeoDet.columns = ['codigoEstacao', 'nivelConsistencia', 'maxima', 'minima','media', 'diaMaxima',
                                   'diaMinima', 'mes', 'ano', 'dataMaxima']
    
    def gerarDfMeo(self):
        try:
            self.dfMeo = pd.DataFrame()
            self.dfMeo['tipoDados'] = self.tipoDados
            self.dfMeo['codEstacao'] = self.codEstacao
            self.dfMeo['nivel'] = self.nivel
            self.dfMeo['totalRegSerie'] = self.totalRegSerie
            self.dfMeo['totalRegSerieNivel'] = self.totalRegSerieNivel
            self.dfMeo['numAnos'] = self.numAnos
            self.dfMeo['numAnosUnico'] = self.numAnosUnico
            self.dfMeo['maiorData'] = self.maiorData
            self.dfMeo['menorData'] = self.menorData
            self.dfMeo['totalRegMaximasAnuais'] = self.totalRegMaximasAnuais
            self.dfMeo['totalRegMaximasAnuaisOrdenado'] = self.totalRegMaximasAnuaisOrdenado
            self.dfMeo['resultado'] = self.resultado
    
            dirInicial = os.path.dirname(__file__).replace('/', '\\')
            #arquivo = f"D:/Desenvolvimento/QGIS 3.34.4/apps/qgis-ltr/python/plugins/pymeo/dados/calcmeo_nivel{self.nivel}_{self.codEstacao}.csv"
            #self.dfMeo.to_csv(arquivo, sep=';', header=True, )
    
            #self.dfMeoMaximasAnuais = self.dfMaximasAnuais_ordenado
            #arquivo = f"D:/Desenvolvimento/QGIS 3.34.4/apps/qgis-ltr/python/plugins/pymeo/dados/calcmeo_maximasanuais_nivel{self.nivel}_{self.codEstacao}.csv"
            #self.dfMeoMaximasAnuais.to_csv(arquivo, sep=';', header=True, )
    
            self.dfMeoSeries = self.df
            #arquivo = 'D:/Desenvolvimento/QGIS 3.34.4/apps/qgis-ltr/python/plugins/pymeo/dados/calcMeo_dfMeoSeries.csv'
            #self.dfMeoSeries.to_csv(arquivo, sep=';', header=True, )
            #print(self.dfMeo)
        except Exception as e:
            self.resultado = -6
            erro = str(e)
            utilsPymeo.gerarLog(2,
                            [f"### Linha: {'{:>4}'.format(inspect.getframeinfo(inspect.currentframe()).lineno)}",
                            f"Problema na execução da função calcular",
                            #f"Status: {self.objHidro.status}",
                            #f"Estação: {self.codEstacao}",
                            #f"Registros no Df: {len(self.objHidro.df)}",
                            erro
                            ], 
                            inspect.getframeinfo(inspect.currentframe()), 
                            self.exibirMsg)
            #sys.exit(1)


#objMeo = meo()
#listaEstacao = ["10100000"]
##
#resultado = []
##
#for codEstacao in listaEstacao:
#    print(datetime.now())
#    objMeo.calcularMeo(codEstacao, origem="Teste calcMeo.py")
#    print(codEstacao, objMeo.resultado)
#    resultado.append((codEstacao, objMeo.resultado))
#    print(datetime.now())